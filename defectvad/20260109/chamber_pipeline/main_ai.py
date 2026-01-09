import os
import sys
import numpy as np
from PIL import Image
import matplotlib.pyplot as plt
from tqdm import tqdm
import pandas as pd
import torch
import gc
import traceback
import json
import re
from torch.utils.data import Dataset, DataLoader
from torchvision import transforms as T
from sklearn.model_selection import train_test_split
import time
from datetime import datetime
from models.components.trainer import EarlyStopper
import platform
import subprocess

# 추가 모듈: Excel 이미지 삽입용
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, Border, Side
from io import BytesIO

# 출력 인코딩 강제 설정 (Windows에서 유용)
import io
import sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# 타임스탬프 생성
TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")

# 스크립트 디렉터리 설정
script_dir = os.path.dirname(os.path.abspath(__file__))
if script_dir not in sys.path:
    sys.path.insert(0, script_dir)

# 결과 저장 디렉터리 설정
RESULTS_SUBDIR = os.path.join(script_dir, "_Results", f"run_{TIMESTAMP}")
os.makedirs(RESULTS_SUBDIR, exist_ok=True)

# _Anomaly 및 _Normal 폴더 생성
ANOMALY_DIR = os.path.join(RESULTS_SUBDIR, "_Anomaly")
NORMAL_DIR = os.path.join(RESULTS_SUBDIR, "_Normal")
os.makedirs(ANOMALY_DIR, exist_ok=True)
os.makedirs(NORMAL_DIR, exist_ok=True)

# 결과 파일 경로
TOTAL_SCORE_CSV_PATH = os.path.join(RESULTS_SUBDIR, f"Total_anomaly_score_{TIMESTAMP}.csv")
TOTAL_SCORE_XLSX_PATH = TOTAL_SCORE_CSV_PATH.replace(".csv", ".xlsx")
LOG_PATH = os.path.join(RESULTS_SUBDIR, f"Inference_Log_List_{TIMESTAMP}.csv")

# 임시 결과 저장 리스트
_temp_results = []

# ----------------------------------------------------------------------
# 입력 데이터 경로 및 색좌표 설정 (메인에서 전달)
# ----------------------------------------------------------------------
data_dir = sys.argv[1] if len(sys.argv) > 1 else os.path.join(script_dir, "data_npz")
output_rgb_dir = os.path.join(data_dir, "data_rgb")

# 외부에서 전달된 색좌표
if len(sys.argv) > 2:
    try:
        primaries = json.loads(sys.argv[2])
        print(f" > primaries : {primaries}")
    except json.JSONDecodeError as e:
        print(f"!!! JSON 파싱 실패: {e}")
        primaries = {
            "W": (0.305, 0.321),
            "R": (0.683, 0.316),
            "G": (0.259, 0.703),
            "B": (0.140, 0.047),
        }
else:
    primaries = {
        "W": (0.305, 0.321),
        "R": (0.683, 0.316),
        "G": (0.259, 0.703),
        "B": (0.140, 0.047),
    }

PROJECT_ROOT = r"E:\_AI_"
TRAINED_DIR = os.path.join(PROJECT_ROOT, "_Trained_models")
BACKBONE_DIR = os.path.join(PROJECT_ROOT, "_Backbones")
DATASET_DIR = data_dir

SEED = 42
NUM_WORKERS = 0
PIN_MEMORY = False
PERSISTENT_WORKERS = False
normalize = True
rotation = 0

# ----------------------------------------------------------------------
#  모델 레지스트리
# ----------------------------------------------------------------------
class ModelRegistry:
    _registry = {}

    @classmethod
    def register(cls, model_type, trainer_path, model_config, train_config):
        cls._registry[model_type] = {
            "trainer_path": trainer_path,
            "model_config": model_config,
            "train_config": train_config
        }

    @classmethod
    def get(cls, model_type):
        if model_type not in cls._registry:
            available = ', '.join(cls.list_models())
            raise ValueError(f"Unknown model_type: '{model_type}'. Available: {available}")
        return cls._registry[model_type]

    @classmethod
    def is_registered(cls, model_type: str) -> bool:
        return model_type in cls._registry

    @classmethod
    def list_models(cls):
        return sorted(cls._registry.keys())

#  get_dataset_dir 함수를 register_all_models() 이전에 정의
def get_dataset_dir():
    return DATASET_DIR

def register_all_models():
    ModelRegistry.register("padim", "models.model_padim.PadimTrainer",
        dict(backbone="wide_resnet50_2", layers=["layer1", "layer2", "layer3"]),
        dict(num_epochs=1, batch_size=4, normalize=True, img_size=256)
    )
    ModelRegistry.register("patchcore", "models.model_patchcore.PatchcoreTrainer",
        dict(backbone="wide_resnet50_2", layers=["layer2", "layer3"]),
        dict(num_epochs=1, batch_size=8, normalize=True, img_size=256)
    )
    ModelRegistry.register("dfkde", "models.model_dfkde.DFKDETrainer",
        dict(backbone="resnet50", layers=["layer4"], pre_trained=True),
        dict(num_epochs=1, batch_size=8, normalize=True, img_size=256)
    )
    ModelRegistry.register("cflow", "models.model_cflow.CflowTrainer",
        dict(backbone="resnet18", layers=["layer2", "layer3"]),
        dict(num_epochs=10, batch_size=4, normalize=True, img_size=256)
    )
    ModelRegistry.register("cflow-resnet18", "models.model_cflow.CflowTrainer",
        dict(backbone="resnet18", layers=["layer2", "layer3"]),
        dict(num_epochs=10, batch_size=4, normalize=True, img_size=256)
    )
    ModelRegistry.register("cflow-resnet50", "models.model_cflow.CflowTrainer",
        dict(backbone="resnet50", layers=["layer2", "layer3"]),
        dict(num_epochs=10, batch_size=4, normalize=True, img_size=256)
    )
    ModelRegistry.register("fastflow", "models.model_fastflow.FastflowTrainer",
        dict(backbone="wide_resnet50_2"),
        dict(num_epochs=20, batch_size=8, normalize=True, img_size=256)
    )
    ModelRegistry.register("fastflow-cait", "models.model_fastflow.FastflowTrainer",
        dict(backbone="cait_m48_448"),
        dict(num_epochs=5, batch_size=4, normalize=True, img_size=448)
    )
    ModelRegistry.register("fastflow-deit", "models.model_fastflow.FastflowTrainer",
        dict(backbone="deit_base_distilled_patch16_384"),
        dict(num_epochs=10, batch_size=8, normalize=True, img_size=384)
    )
    ModelRegistry.register("csflow", "models.model_csflow.CsFlowTrainer",
        dict(num_channels=3),
        dict(num_epochs=10, batch_size=8, normalize=True, img_size=256)
    )
    ModelRegistry.register("uflow", "models.model_uflow.UflowTrainer",
        dict(backbone="wide_resnet50_2"),
        dict(num_epochs=10, batch_size=8, normalize=True, img_size=256)
    )
    ModelRegistry.register("uflow-mcait", "models.model_uflow.UflowTrainer",
        dict(backbone="mcait"),
        dict(num_epochs=10, batch_size=4, normalize=True, img_size=448)
    )
    ModelRegistry.register("stfpm", "models.model_stfpm.STFPMTrainer",
        dict(backbone="resnet50", layers=["layer1", "layer2", "layer3"]),
        dict(num_epochs=50, batch_size=16, normalize=True, img_size=256)
    )
    ModelRegistry.register("fre", "models.model_fre.FRETrainer",
        dict(backbone="resnet50", layer="layer3"),
        dict(num_epochs=50, batch_size=16, normalize=True, img_size=256)
    )
    ModelRegistry.register("reverse-distillation", "models.model_reverse_distillation.ReverseDistillationTrainer",
        dict(backbone="wide_resnet50_2", layers=["layer1", "layer2", "layer3"]),
        dict(num_epochs=50, batch_size=8, normalize=True, img_size=256)
    )
    ModelRegistry.register("efficientad-small", "models.model_efficientad.EfficientAdTrainer",
        dict(model_size="small", imagenet_dir=os.path.join(get_dataset_dir(), "imagenette2")),
        dict(num_epochs=20, batch_size=1, normalize=False, img_size=256)
    )
    ModelRegistry.register("efficientad-medium", "models.model_efficientad.EfficientAdTrainer",
        dict(model_size="medium", imagenet_dir=os.path.join(get_dataset_dir(), "imagenette2")),
        dict(num_epochs=20, batch_size=1, normalize=False, img_size=256)
    )
    ModelRegistry.register("efficientad", "models.model_efficientad.EfficientAdTrainer",
        dict(model_size="medium", imagenet_dir=os.path.join("/home/namu/myspace/NAMU/backbones", "imagenette2"),
            early_stopper_auroc=EarlyStopper(target_value=0.998)),
        dict(num_epochs=50, batch_size=1, normalize=False, img_size=256)
    )
    ModelRegistry.register("autoencoder", "models.model_autoencoder.AutoencoderTrainer",
        dict(latent_dim=128),
        dict(num_epochs=50, batch_size=16, normalize=False, img_size=256)
    )
    ModelRegistry.register("ganomaly", "models.model_ganomaly.GanomalyTrainer",
        dict(input_size=(256, 256), n_features=64, latent_vec_size=256, gamma=0.01),
        dict(num_epochs=20, batch_size=8, normalize=False, img_size=256)
    )
    ModelRegistry.register("draem", "models.model_draem.DraemTrainer",
        dict(sspcab=True, dtd_dir=os.path.join(get_dataset_dir(), "dtd")),
        dict(num_epochs=10, batch_size=8, normalize=False, img_size=256)
    )
    ModelRegistry.register("dsr", "models.model_dsr.DsrTrainer",
        dict(latent_anomaly_strength=0.2, embedding_dim=128, num_embeddings=4096),
        dict(num_epochs=50, batch_size=8, normalize=False, img_size=256)
    )
    ModelRegistry.register("dfm", "models.model_dfm.DFMTrainer",
        dict(backbone="resnet50", layer="layer3", score_type="fre"),
        dict(num_epochs=1, batch_size=16, normalize=True, img_size=256)
    )
    ModelRegistry.register("cfa", "models.model_cfa.CfaTrainer",
        dict(backbone="wide_resnet50_2"),
        dict(num_epochs=20, batch_size=16, normalize=True, img_size=256)
    )
    ModelRegistry.register("dinomaly-small-224", "models.model_dinomaly.DinomalyTrainer",
        dict(encoder_name="dinov2_vit_small_14", bottleneck_dropout=0.2, decoder_depth=8),
        dict(num_epochs=15, batch_size=32, normalize=True, img_size=224)
    )
    ModelRegistry.register("dinomaly-base-224", "models.model_dinomaly.DinomalyTrainer",
        dict(encoder_name="dinov2_vit_base_14", bottleneck_dropout=0.2, decoder_depth=8),
        dict(num_epochs=15, batch_size=16, normalize=True, img_size=224)
    )
    ModelRegistry.register("dinomaly-large-224", "models.model_dinomaly.DinomalyTrainer",
        dict(encoder_name="dinov2_vit_large_14", bottleneck_dropout=0.2, decoder_depth=8),
        dict(num_epochs=15, batch_size=8, normalize=True, img_size=224)
    )
    ModelRegistry.register("dinomaly-small-392", "models.model_dinomaly.DinomalyTrainer",
        dict(encoder_name="dinov2_vit_small_14", bottleneck_dropout=0.2, decoder_depth=8),
        dict(num_epochs=10, batch_size=24, normalize=True, img_size=392)
    )
    ModelRegistry.register("dinomaly-base-392", "models.model_dinomaly.DinomalyTrainer",
        dict(encoder_name="dinov2_vit_base_14", bottleneck_dropout=0.2, decoder_depth=8),
        dict(num_epochs=10, batch_size=12, normalize=True, img_size=392)
    )
    ModelRegistry.register("dinomaly-large-392", "models.model_dinomaly.DinomalyTrainer",
        dict(encoder_name="dinov2_vit_large_14", bottleneck_dropout=0.2, decoder_depth=8),
        dict(num_epochs=10, batch_size=6, normalize=True, img_size=392)
    )
    ModelRegistry.register("dinomaly-small-448", "models.model_dinomaly.DinomalyTrainer",
        dict(encoder_name="dinov2_vit_small_14", bottleneck_dropout=0.2, decoder_depth=8),
        dict(num_epochs=10, batch_size=16, normalize=True, img_size=448)
    )
    ModelRegistry.register("dinomaly-base-448", "models.model_dinomaly.DinomalyTrainer",
        dict(encoder_name="dinov2_vit_base_14", bottleneck_dropout=0.2, decoder_depth=8),
        dict(num_epochs=10, batch_size=8, normalize=True, img_size=448)
    )
    ModelRegistry.register("dinomaly-large-448", "models.model_dinomaly.DinomalyTrainer",
        dict(encoder_name="dinov2_vit_large_14", bottleneck_dropout=0.2, decoder_depth=8),
        dict(num_epochs=10, batch_size=4, normalize=True, img_size=448)
    )
    ModelRegistry.register("supersimplenet", "models.model_supersimplenet.SupersimplenetTrainer",
        dict(backbone="wide_resnet50_2", layers=["layer2", "layer3"], supervised=False),
        dict(num_epochs=50, batch_size=16, normalize=True, img_size=256)
    )
    ModelRegistry.register("supersimplenet-supervised", "models.model_supersimplenet.SupersimplenetTrainer",
        dict(backbone="wide_resnet50_2", layers=["layer2", "layer3"], supervised=True),
        dict(num_epochs=50, batch_size=16, normalize=True, img_size=256)
    )
    ModelRegistry.register("uninet", "models.model_uninet.UniNetTrainer",
        dict(student_backbone="wide_resnet50_2", teacher_backbone="wide_resnet50_2", temperature=0.4),
        dict(num_epochs=20, batch_size=4, normalize=True, img_size=256)
    )

register_all_models()

# --- 수학 함수 정의 ---
def xy_to_XYZ(x, y, Y=1.0):
    X = x * (Y / y)
    Z = (1 - x - y) * (Y / y)
    return np.array([X, Y, Z], dtype=np.float32)

def get_RGB2XYZ_matrix(primaries, Y_white):
    XYZ_r = xy_to_XYZ(*primaries["R"], Y=1.0)
    XYZ_g = xy_to_XYZ(*primaries["G"], Y=1.0)
    XYZ_b = xy_to_XYZ(*primaries["B"], Y=1.0)
    XYZ_w = xy_to_XYZ(*primaries["W"], Y=Y_white)
    matrix = np.stack([XYZ_r, XYZ_g, XYZ_b], axis=-1).astype(np.float32)
    scale = np.linalg.solve(matrix, XYZ_w).astype(np.float32)
    return matrix * scale[np.newaxis, :]

def linear_to_srgb(linear):
    linear = np.clip(linear, 0.0, 1.0)
    mask = linear <= 0.0031308
    srgb = np.empty_like(linear, dtype=np.float32)
    srgb[mask] = linear[mask] * 12.92
    srgb[~mask] = 1.055 * np.power(linear[~mask], 1.0/2.4) - 0.055
    return srgb

def XYZ_to_RGB(XYZ, primaries, Y_white):
    M_RGB2XYZ = get_RGB2XYZ_matrix(primaries, Y_white)
    M_XYZ2RGB = np.linalg.inv(M_RGB2XYZ)
    linear = M_XYZ2RGB @ XYZ.reshape(-1, 3).T
    linear = linear.T.reshape(XYZ.shape)
    linear = np.clip(linear, 0, 1).astype(np.float32)
    return linear_to_srgb(linear)

def get_XYZ_from_npz(path, rotation=0):
    if not os.path.exists(path):
        raise FileNotFoundError(f"NPZ 파일 없음: {path}")
    XYZ = np.load(path)["data"]
    if rotation == 180:
        XYZ = np.flip(XYZ, axis=(0, 1))
    return XYZ

def get_filenames_dimmings(data_dir):
    filenames = []
    dimmings = []
    if not os.path.exists(data_dir):
        return filenames, dimmings
    for filename in sorted(os.listdir(data_dir)):
        if filename.endswith("_f16.npz"):
            filenames.append(filename)
            try:
                dimming = filename.split('.')[0].split()[-1].split('_')[0]
            except:
                dimming = "1"
            dimmings.append(dimming)
    return filenames, dimmings

def save_npz_to_rgb(data_dir, output_rgb_dir, rotation=0, primaries=None, normalize=True):
    if primaries is None:
        primaries = {
            "W": (0.3127, 0.3290),
            "R": (0.640, 0.330),
            "G": (0.300, 0.600),
            "B": (0.150, 0.060),
        }
    if not os.path.exists(output_rgb_dir):
        os.makedirs(output_rgb_dir)

    if not os.path.exists(data_dir):
        raise FileNotFoundError(f"데이터 디렉터리 없음: {data_dir}")

    filenames, dimmings = get_filenames_dimmings(data_dir)
    if len(filenames) == 0:
        return

    print(f"[2/8] 📦 RGB 이미지 변환 시작 | {len(filenames)}개의 NPZ 파일 발견")
    total = len(filenames)
    num_images = 0

    with tqdm(zip(filenames, dimmings), desc="RGB 이미지 저장 중", total=total, ascii=True, leave=False) as pbar:
        for filename, dimming in pbar:
            try:
                image_name = filename.replace('_f16.npz', '_rgb.png')
                pbar.set_postfix_str(image_name)

                image_path = os.path.join(data_dir, filename)
                XYZ = get_XYZ_from_npz(image_path, rotation=rotation)
                XYZ = XYZ / float(dimming)

                if normalize:
                    Y = XYZ[..., 1]
                    min_val, max_val = np.min(Y), np.max(Y)
                    if max_val > min_val:
                        XYZ = (XYZ - min_val) / (max_val - min_val)

                RGB_linear = XYZ_to_RGB(XYZ, primaries=primaries, Y_white=1.0)

                for c in range(3):
                    channel = RGB_linear[..., c]
                    min_val, max_val = channel.min(), channel.max()
                    if max_val > min_val:
                        RGB_linear[..., c] = (channel - min_val) / (max_val - min_val)

                RGB_srgb = np.clip(RGB_linear * 255.0, 0, 255).astype('uint8')
                image = Image.fromarray(RGB_srgb, 'RGB')
                image.save(os.path.join(output_rgb_dir, image_name))
                num_images += 1
            except Exception as e:
                continue

    print(f"[2/8] ✅ RGB 이미지 변환 완료 | {num_images}개 저장됨: {output_rgb_dir}")

# ----------------------------------------------------------------------
# Custom Dataset
# ----------------------------------------------------------------------
def get_data_info(image_path):
    basename = os.path.basename(image_path)
    filename_no_ext = os.path.splitext(basename)[0]
    if filename_no_ext.endswith('_rgb'):
        filename_no_ext = filename_no_ext[:-4]
    parts = filename_no_ext.split()
    if len(parts) < 1:
        raise ValueError(f"Invalid filename: {image_path}")
    category_raw = parts[0]
    category = category_raw.rstrip('_')
    info = {
        "filename": basename,
        "category": category,
        "freq": float(''.join(filter(str.isdigit, parts[1]))) if len(parts) > 1 else 0.0,
        "dimming": float(''.join(filter(str.isdigit, parts[2]))) if len(parts) > 2 else 1.0,
        "image_path": image_path,
        "dataset_type": "custom",
        "defect_type": "normal",
        "label": 0
    }
    return info

def create_csv(csv_path):
    data_dir = os.path.dirname(csv_path)
    if not os.path.exists(data_dir):
        os.makedirs(data_dir)
    records = []
    for filename in sorted(os.listdir(data_dir)):
        if not (filename.endswith(".png") and "_rgb.png" in filename):
            continue
        image_path = os.path.join(data_dir, filename)
        if not os.path.exists(image_path):
            continue
        try:
            info = get_data_info(image_path)
            records.append(info)
        except Exception as e:
            continue
    if not records:
        raise ValueError(f"CSV 생성 실패: 유효한 이미지 없음 - {data_dir}")
    df = pd.DataFrame(records)
    df.to_csv(csv_path, index=False)

class CustomDataset(Dataset):
    def __init__(self, dataset_dir, category, split="test", transform=None, mask_transform=None, test_ratio=0.2):
        super().__init__()
        self.transform = transform
        self.mask_transform = mask_transform
        self.image_paths = []
        self.labels = []
        self.defect_types = []
        self.categories = []
        self.has_mask = False
        self.load_data(dataset_dir, category, split, test_ratio)

    def load_data(self, dataset_dir, category, split, test_ratio):
        data_dir = os.path.join(dataset_dir, "data_rgb")
        csv_path = os.path.join(data_dir, "data_info.csv")
        if not os.path.exists(csv_path):
            create_csv(csv_path)
        df = pd.read_csv(csv_path)
        df['category_clean'] = df['category'].str.rstrip('_')
        category_clean = category.rstrip('_')

        if category != "all":
            df = df[df["category_clean"] == category_clean].reset_index(drop=True)
            if len(df) == 0:
                return

        normal_df = df[df["label"] == 0].reset_index(drop=True)
        if len(normal_df) == 0:
            return

        if len(normal_df) == 1:
            test_normal = normal_df.copy()
        else:
            _, test_normal = train_test_split(normal_df, test_size=test_ratio, random_state=42, shuffle=True)

        subset = test_normal.reset_index(drop=True)
        valid_count = 0
        for _, row in subset.iterrows():
            image_path = row["image_path"]
            if not os.path.exists(image_path):
                continue
            self.image_paths.append(image_path)
            self.labels.append(int(row["label"]))
            self.defect_types.append(row["defect_type"])
            self.categories.append(row["category"])
            valid_count += 1

    def __len__(self):
        return len(self.image_paths)

    def __getitem__(self, idx):
        try:
            image_path = self.image_paths[idx]
            image = Image.open(image_path).convert('RGB')
            if self.transform:
                image = self.transform(image)
        except Exception as e:
            image = torch.zeros(3, 224, 224, dtype=torch.float32)

        label = torch.tensor(self.labels[idx]).long()
        name = os.path.basename(self.image_paths[idx]) if idx < len(self.image_paths) else f"unknown_{idx}"
        defect_type = self.defect_types[idx] if idx < len(self.defect_types) else "normal"
        category = self.categories[idx] if idx < len(self.categories) else "unknown"

        height, width = image.shape[-2:] if isinstance(image, torch.Tensor) else (224, 224)
        mask = torch.zeros((height, width), dtype=torch.long)

        return dict(
            image=image,
            label=label,
            mask=mask,
            name=name,
            category=category,
            defect_type=defect_type,
            has_mask=False,
            image_path=image_path
        )

# ----------------------------------------------------------------------
# Data Loader
# ----------------------------------------------------------------------
_loader_cache = {}
def get_test_loader(dataset_dir, dataset_type, category, img_size, batch_size, normalize=True,
                    test_ratio=0.2, num_workers=8, pin_memory=True, persistent_workers=False):
    cache_key = (category, img_size, normalize)
    if cache_key in _loader_cache:
        return _loader_cache[cache_key]

    test_transforms = [
        T.Resize((img_size, img_size), interpolation=T.InterpolationMode.BILINEAR),
        T.ToTensor(),
    ]
    if normalize:
        normalize_transform = T.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225])
        test_transforms.append(normalize_transform)
    test_transform = T.Compose(test_transforms)

    dataset = CustomDataset(dataset_dir, category=category, split="test", transform=test_transform, test_ratio=test_ratio)
    if len(dataset) == 0:
        return None
    loader = DataLoader(dataset, batch_size=batch_size, shuffle=False, num_workers=num_workers,
                        pin_memory=pin_memory, persistent_workers=persistent_workers)
    _loader_cache[cache_key] = loader
    return loader

# ----------------------------------------------------------------------
# 유틸리티
# ----------------------------------------------------------------------
def set_globals(dataset_dir=None, backbone_dir=None, seed=None, num_workers=None,
                pin_memory=None, persistent_workers=None, show_globals=False):
    global DATASET_DIR, BACKBONE_DIR, SEED, NUM_WORKERS, PIN_MEMORY, PERSISTENT_WORKERS
    if dataset_dir is not None: DATASET_DIR = dataset_dir
    if seed is not None: SEED = seed
    if num_workers is not None: NUM_WORKERS = num_workers
    if pin_memory is not None: PIN_MEMORY = pin_memory
    if persistent_workers is not None: PERSISTENT_WORKERS = persistent_workers
    if backbone_dir is not None:
        BACKBONE_DIR = backbone_dir
        try:
            from models.components.backbone import set_backbone_dir
            set_backbone_dir(BACKBONE_DIR)
        except Exception as e:
            pass
    if show_globals: print_globals()

def print_globals():
    cfg = {"dataset_dir": DATASET_DIR, "backbone_dir": BACKBONE_DIR,
           "seed": SEED, "num_workers": NUM_WORKERS, "pin_memory": PIN_MEMORY, "persistent_workers": PERSISTENT_WORKERS}
    print("\n" + "=" * 70)
    print("⚙️ 추론 설정 정보")
    print("-" * 70)
    for k, v in cfg.items(): print(f"  {k:20s}: {v}")
    print("-" * 70)

def set_seed(seed: int = 42):
    import random, numpy as np
    random.seed(seed); np.random.seed(seed); torch.manual_seed(seed); torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True; torch.backends.cudnn.benchmark = False
    os.environ["PYTHONHASHSEED"] = str(seed)

def clear_memory(print_summary: bool = True, stage: str = "정리 후"):
    collected = gc.collect()
    if torch.cuda.is_available():
        torch.cuda.empty_cache(); torch.cuda.synchronize()

def print_memory(stage: str = ""):
    pass

# ----------------------------------------------------------------------
# 결과 시각화
# ----------------------------------------------------------------------
def normalize_map(anomaly_map, method='percentile'):
    if method == 'minmax':
        min_val, max_val = anomaly_map.min(), anomaly_map.max()
        return (anomaly_map - min_val) / (max_val - min_val + 1e-8)
    elif method == 'percentile':
        p1, p99 = np.percentile(anomaly_map, [1, 99])
        anomaly_map = np.clip(anomaly_map, p1, p99)
        return (anomaly_map - p1) / (p99 - p1 + 1e-8)
    else:
        return anomaly_map

def create_border_mask(H, W, border_ratio=0.1):
    mask = np.ones((H, W), dtype=np.float32)
    t = int(H * border_ratio)
    l = int(W * border_ratio)
    b = H - t
    r = W - l
    mask[t:b, l:r] = 0
    return mask

def visualize_triple_plot(image_path, normalized_image, anomaly_map, output_path, anomaly_score, filename, threshold, pred_label, method, task, raw_img):
    H, W, _ = raw_img.shape
    FIXED_HEIGHT = 500
    dpi = 100
    scale = FIXED_HEIGHT / H
    new_width = int(W * scale)
    figsize = (3 * new_width / dpi, FIXED_HEIGHT / dpi)

    fig, axes = plt.subplots(1, 3, figsize=figsize, dpi=dpi)
    title_text = f"\n{filename}  |  Anomaly Score: {anomaly_score:.6f}\n\n" \
                 f"Model: {method} | Category: {task} | Threshold: {threshold:.6f} => {pred_label.upper()}\n"
    fig.suptitle(title_text, fontsize=14, fontweight='bold', y=0.97, ha='center', va='top')

    axes[0].imshow(raw_img); axes[0].set_title("Raw Image", fontsize=11); axes[0].axis('off')
    axes[1].imshow(normalized_image); axes[1].set_title("Normalized RGB", fontsize=11); axes[1].axis('off')

    border_mask = create_border_mask(H, W, border_ratio=0.1)
    masked_map = np.where(border_mask == 1, 0, anomaly_map)
    axes[2].imshow(masked_map, cmap='jet'); axes[2].set_title("Anomaly Map (Outer 10% masked)", fontsize=11); axes[2].axis('off')

    plt.subplots_adjust(left=0.05, right=0.95, top=0.70, bottom=0.10, wspace=0.57)
    plt.savefig(output_path, bbox_inches='tight', pad_inches=0.10, dpi=dpi)
    plt.close()

# ----------------------------------------------------------------------
# 결과 저장
# ----------------------------------------------------------------------
def save_anomaly_map_with_raw(anomaly_map, method, task, raw_image_path, anomaly_score, filename, threshold, pred_label):
    save_dir = ANOMALY_DIR if pred_label == "anomaly" else NORMAL_DIR

    if isinstance(filename, (list, tuple)): filename = filename[0]
    if isinstance(raw_image_path, (list, tuple)): raw_image_path = raw_image_path[0]
    if not raw_image_path or not os.path.exists(raw_image_path):
        return

    try:
        raw_img = np.array(Image.open(raw_image_path).convert('RGB'))
        H, W = raw_img.shape[:2]
    except Exception as e:
        H, W = 224, 224
        raw_img = np.zeros((H, W, 3), dtype=np.uint8)

    from skimage.transform import resize
    anomaly_map_resized = resize(anomaly_map, (H, W), preserve_range=True, anti_aliasing=True)
    norm_map = normalize_map(anomaly_map_resized, method='percentile')
    norm_img = np.clip((raw_img.astype(np.float32) / 255.0 - [0.485, 0.456, 0.406]) / [0.229, 0.224, 0.225], 0, 1)

    border_mask = create_border_mask(H, W, border_ratio=0.1)
    inner_mask = border_mask == 0
    inner_map = anomaly_map_resized[inner_mask]
    anomaly_score = float(np.mean(inner_map)) if inner_map.size > 0 else float(np.mean(anomaly_map_resized))

    base_filename = os.path.splitext(os.path.basename(raw_image_path))[0]
    output_filename = f"Anomaly_Map_{pred_label}_{method}_{base_filename}.png"
    output_path = os.path.join(save_dir, output_filename)
    visualize_triple_plot(raw_image_path, norm_img, norm_map, output_path, anomaly_score, filename,
                          threshold, pred_label, method, task, raw_img)

def save_to_total_csv(filename, method, task, anomaly_score, threshold, pred_label, processing_time):
    record = {
        "Filename": filename,
        "Model": method,
        "Category": task,
        "Anomaly Score": anomaly_score,
        "Threshold": threshold,
        "Label": pred_label,
        "Processing Time (s)": f"{processing_time:.4f}"
    }
    _temp_results.append(record)

def flush_results():
    if not _temp_results:
        return

    df = pd.DataFrame(_temp_results)
    df.insert(0, "Index", range(1, len(df) + 1))
    df.to_csv(TOTAL_SCORE_CSV_PATH, index=False, encoding='utf-8-sig')

    with pd.ExcelWriter(TOTAL_SCORE_XLSX_PATH, engine='openpyxl') as writer:
        df_normal = df[df["Label"] == "normal"].reset_index(drop=True)
        df_anomaly = df[df["Label"] == "anomaly"].reset_index(drop=True)

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # Normal 시트
        if not df_normal.empty:
            df_normal.to_excel(writer, index=False, sheet_name='Normal')
            worksheet = writer.sheets['Normal']
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
            worksheet.row_dimensions[1].height = 22.5  # 1.5배

            img_col = len(df_normal.columns) + 1
            cell_header = worksheet.cell(row=1, column=img_col, value="Debug Image")
            cell_header.font = Font(bold=True)
            cell_header.alignment = Alignment(horizontal='center', vertical='center')
            cell_header.border = thin_border

            col_letter = openpyxl.utils.get_column_letter(img_col)
            image_width_px = 750
            image_height_px = 300
            worksheet.column_dimensions[col_letter].width = (image_width_px * 1.05) / 7  # 5% 더 넓게

            for idx, row in df_normal.iterrows():
                base_filename = os.path.splitext(row["Filename"])[0]
                img_path = os.path.join(NORMAL_DIR, f"Anomaly_Map_normal_{row['Model']}_{base_filename}.png")
                if os.path.exists(img_path):
                    img = XLImage(img_path)
                    img.width = image_width_px
                    img.height = image_height_px
                    worksheet.add_image(img, f"{col_letter}{idx + 2}")
                    worksheet.row_dimensions[idx + 2].height = (image_height_px * 1.05) * 0.75  # 5% 더 높게

        # Anomaly 시트
        if not df_anomaly.empty:
            df_anomaly.to_excel(writer, index=False, sheet_name='Anomaly')
            worksheet = writer.sheets['Anomaly']
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
            worksheet.row_dimensions[1].height = 22.5

            img_col = len(df_anomaly.columns) + 1
            cell_header = worksheet.cell(row=1, column=img_col, value="Debug Image")
            cell_header.font = Font(bold=True)
            cell_header.alignment = Alignment(horizontal='center', vertical='center')
            cell_header.border = thin_border

            col_letter = openpyxl.utils.get_column_letter(img_col)
            worksheet.column_dimensions[col_letter].width = (image_width_px * 1.05) / 7

            for idx, row in df_anomaly.iterrows():
                base_filename = os.path.splitext(row["Filename"])[0]
                img_path = os.path.join(ANOMALY_DIR, f"Anomaly_Map_anomaly_{row['Model']}_{base_filename}.png")
                if os.path.exists(img_path):
                    img = XLImage(img_path)
                    img.width = image_width_px
                    img.height = image_height_px
                    worksheet.add_image(img, f"{col_letter}{idx + 2}")
                    worksheet.row_dimensions[idx + 2].height = (image_height_px * 1.05) * 0.75

        for sheetname in writer.sheets:
            ws = writer.sheets[sheetname]
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

    print(f"    > 📊 최종 결과 저장: {TOTAL_SCORE_XLSX_PATH}")

def get_threshold_from_weight_path(weight_path):
    if not weight_path: return 0.0
    fname = os.path.basename(weight_path)
    if "F1_" in fname:
        try:
            f1_val = float(fname.split("F1_")[1].split("_")[0])
            return f1_val
        except:
            pass
    return 0.0

_weight_cache = {}
def find_weight_path(task: str, method: str):
    cache_key = (task, method)
    if cache_key in _weight_cache:
        return _weight_cache[cache_key]
    if not os.path.exists(TRAINED_DIR):
        _weight_cache[cache_key] = None
        return None
    for fname in os.listdir(TRAINED_DIR):
        if not fname.endswith(".pth"):
            continue
        fpath = os.path.join(TRAINED_DIR, fname)
        if os.path.getsize(fpath) == 0:
            continue
        task_clean = task.rstrip('_')
        fname_clean = fname.replace('_', '')
        if task_clean.replace('_', '') in fname_clean and method in fname:
            _weight_cache[cache_key] = fpath
            return fpath
    _weight_cache[cache_key] = None
    return None

_trainer_cache = {}
def get_trainer(model_type, img_size):
    cache_key = model_type
    if cache_key in _trainer_cache:
        return _trainer_cache[cache_key]
    if not ModelRegistry.is_registered(model_type):
        raise ValueError(f"Unknown model_type: '{model_type}'")
    config = ModelRegistry.get(model_type)
    module_path, class_name = config["trainer_path"].rsplit(".", 1)
    module = __import__(module_path, fromlist=[class_name])
    TrainerClass = getattr(module, class_name)
    model_config = config["model_config"].copy()
    trainer = TrainerClass(**model_config)
    _trainer_cache[cache_key] = trainer
    return trainer

# ----------------------------------------------------------------------
# 추론 실행
# ----------------------------------------------------------------------
def load_trainer_and_weights(model_type, img_size, weight_path):
    try:
        trainer = get_trainer(model_type, img_size=img_size)
        if not hasattr(trainer, '_loaded_weights') or trainer._loaded_weights != weight_path:
            load_start = time.time()
            checkpoint = torch.load(weight_path, map_location='cpu')
            model_state = checkpoint.get("model", checkpoint)
            model_dict = trainer.model.state_dict()
            filtered_state = {}
            for k, v in model_state.items():
                if k in model_dict and v.shape == model_dict[k].shape:
                    filtered_state[k] = v
            trainer.model.load_state_dict(filtered_state, strict=False)
            trainer._loaded_weights = weight_path
            load_end = time.time()

        device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        trainer.model.to(device)
        trainer.model.eval()
        return trainer
    except Exception as e:
        traceback.print_exc()
        raise e

def run_inference(category, model_type, img_size=None, normalize=None):
    if not ModelRegistry.is_registered(model_type):
        return False

    train_cfg = get_train_config(model_type)
    img_size = img_size or train_cfg["img_size"]
    normalize = normalize if normalize is not None else train_cfg.get("normalize", True)

    weight_path = find_weight_path(task=category, method=model_type)
    if not weight_path:
        return False

    set_seed(SEED)

    test_loader = get_test_loader(
        dataset_dir=DATASET_DIR,
        dataset_type="custom",
        category=category,
        img_size=img_size,
        batch_size=1,
        normalize=normalize,
        num_workers=NUM_WORKERS,
        pin_memory=PIN_MEMORY,
        persistent_workers=PERSISTENT_WORKERS
    )
    if test_loader is None or len(test_loader) == 0:
        return False

    try:
        trainer = load_trainer_and_weights(model_type, img_size, weight_path)
    except Exception as e:
        return False

    threshold = get_threshold_from_weight_path(weight_path)

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    total_images = len(test_loader)
    print(f"Total images to process: {total_images}")

    try:
        for idx, batch in enumerate(test_loader):
            try:
                img = batch["image"].to(device)
                filename = batch.get("name", [f"image_{idx}"])[0]
                raw_image_path = batch.get("image_path", [None])[0]

                if img.shape[-1] != img_size or img.shape[-2] != img_size:
                    transform_resize = T.Resize((img_size, img_size), interpolation=T.InterpolationMode.BILINEAR)
                    img = transform_resize(img)

                start_time = time.time()
                with torch.no_grad():
                    out = trainer.model(img)
                anomaly_map = out.get("anomaly_map", out.get("pred", list(out.values())[0]))
                anomaly_map = anomaly_map.squeeze().detach().cpu().numpy()

                H, W = img.shape[-2], img.shape[-1]
                border_mask = create_border_mask(H, W, border_ratio=0.1)
                inner_mask = border_mask == 0
                if anomaly_map.shape == (H, W):
                    inner_map = anomaly_map[inner_mask]
                else:
                    inner_map = anomaly_map
                anomaly_score = float(np.mean(inner_map)) if inner_map.size > 0 else float(np.mean(anomaly_map))

                pred_label = "anomaly" if anomaly_score > threshold else "normal"

                save_anomaly_map_with_raw(anomaly_map, method=model_type, task=category,
                                          raw_image_path=raw_image_path, anomaly_score=anomaly_score,
                                          filename=filename, threshold=threshold, pred_label=pred_label)

                end_time = time.time()
                processing_time = end_time - start_time
                save_to_total_csv(filename, model_type, category, anomaly_score, threshold, pred_label, processing_time)

                print(f"[진행 중] [{idx + 1}/{total_images}] {filename}")

            except Exception as e:
                continue
    except KeyboardInterrupt:
        return False

    return True

def get_train_config(model_type):
    config_map = {
        "dinomaly-large-448": {"img_size": 448, "num_epochs": 1, "normalize": True},
        "stfpm": {"img_size": 224, "num_epochs": 1, "normalize": True},
        "efficientad": {"img_size": 256, "num_epochs": 1, "normalize": True},
    }
    return config_map.get(model_type, {"img_size": 256, "num_epochs": 1, "normalize": True})

# ----------------------------------------------------------------------
# 추론 배치 실행
# ----------------------------------------------------------------------
def inference_batch(categories, methods, img_size=None, normalize=None, clear_memory_between=True):
    if isinstance(categories, str): categories = [categories]
    if isinstance(methods, str): methods = [methods]

    valid_runs = []
    for cat in categories:
        for m in methods:
            weight_path = find_weight_path(cat, m)
            if ModelRegistry.is_registered(m) and weight_path:
                valid_runs.append((cat, m))

    if not valid_runs:
        return []

    all_results = []
    for idx, (cat, m) in enumerate(valid_runs, 1):
        print(f"\n[진행 중] [{idx}/{len(valid_runs)}] {m} | {cat}")
        try:
            success = run_inference(cat, m, img_size, normalize)
        except KeyboardInterrupt:
            break
        all_results.append({"category": cat, "model": m, "status": "success" if success else "failed"})
        if clear_memory_between and idx < len(valid_runs):
            clear_memory(print_summary=False)

    return all_results

def _open_file(filepath: str):
    """
    운영체제에 맞는 방법으로 파일을 열어줍니다.
    - Windows : os.startfile
    - macOS   : `open` 커맨드
    - Linux   : `xdg-open` 커맨드
    """
    try:
        if platform.system() == "Windows":
            os.startfile(filepath)                     # type: ignore
        elif platform.system() == "Darwin":           # macOS
            subprocess.run(["open", filepath], check=False)
        else:                                         # Linux, 기타 UNIX 계열
            subprocess.run(["xdg-open", filepath], check=False)
    except Exception as exc:
        print(f"[경고] 엑셀 파일을 자동으로 열 수 없습니다: {exc}")

# ----------------------------------------------------------------------
# 메인 실행
# ----------------------------------------------------------------------
if __name__ == "__main__":
    try:
        set_globals(backbone_dir=BACKBONE_DIR, show_globals=True)

        if not os.path.exists(output_rgb_dir):
            os.makedirs(output_rgb_dir)

        save_npz_to_rgb(data_dir, output_rgb_dir=output_rgb_dir, rotation=rotation, primaries=primaries, normalize=normalize)

        set_globals(seed=SEED, num_workers=NUM_WORKERS, pin_memory=PIN_MEMORY, persistent_workers=False)

        task_groups = [
            "t2_4_d", "t2_4_e", "t2_4_f", "t2_4_g", "t2_4_h", "t2_4_i",
            "t2_10_a", "t2_10_b", "t2_10_c", "t2_10_f", "t2_10_g", "t2_10_i", "t2_10_k",
            "t2_16_c", "t2_16_d", "t2_16_e",
            "t2_18_b", "t2_18_c", "t2_18_e", "t2_18_f", "t2_18_g",
            "VScale_0_255", "VScale_B_0_255", "VScale_G_0_255", "VScale_R_0_255",
            "VScale_255_0", "VScale_B_255_0", "VScale_G_255_0", "VScale_R_255_0",
            "W16", "W24", "W48", "W72", "W164", "W255",
            "4H", "B255", "G255", "R255",
        ]

        methods = ["stfpm", "reverse-distillation", "efficientad", "dinomaly-large-448"]

        total_results = []
        for task in task_groups:
            for method in methods:
                results = inference_batch([task], [method])
                total_results.extend(results)

        success_cnt = sum(1 for r in total_results if r["status"] == "success")
        fail_cnt = sum(1 for r in total_results if r["status"] == "failed")
        total_cnt = len(total_results)
        print(f"\n[8/8] # 최종 추론 요약")
        print(f"    > 전체 작업 수: {total_cnt}")
        print(f"    > 성공: {success_cnt} | 실패: {fail_cnt} | 성공률: {success_cnt / total_cnt * 100:.1f}%")

    except KeyboardInterrupt:
        pass
    except Exception as e:
        traceback.print_exc()
    finally:
        flush_results()

        if os.path.isfile(TOTAL_SCORE_XLSX_PATH):
            print(f"\n # 엑셀 파일을 자동으로 엽니다: {TOTAL_SCORE_XLSX_PATH}")
            _open_file(TOTAL_SCORE_XLSX_PATH)
        else:
            print("\n # 엑셀 파일이 존재하지 않아 열 수 없습니다.")

        print(f"\n # 전체 작업 완료!")
        print(f" # 통합 결과 폴더: {RESULTS_SUBDIR}")
        print(f"   - Anomaly Maps: {os.path.join(RESULTS_SUBDIR, '_Anomaly')}")
        print(f"   - Normal Maps: {os.path.join(RESULTS_SUBDIR, '_Normal')}")
        print(f"   - Total Score XLSX: {TOTAL_SCORE_XLSX_PATH}")
